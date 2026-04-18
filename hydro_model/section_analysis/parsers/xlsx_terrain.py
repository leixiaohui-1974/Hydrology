# ALGORITHM_REGISTRY:
#   id: xlsx_terrain_section_parser
#   category: hydrology
#   protocol: HydrologyProduct
#   case_types: [cascade_hydro, water_transfer]
#   source_project: hydrology
"""实测河道地形 XLSX 断面解析器。

支持格式:
  - 大渡河数据集/断面、地形、雨量、坐标/实测河道地形/*.xlsx
  - 断面坐标1.xlsx
  - 各种以 sheet 分断面或行分断面的 Excel 文件
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from ..base import SectionProfile


class XlsxTerrainParser:
    """从 Excel 文件解析河道断面数据。"""

    @staticmethod
    def can_handle(path: str) -> bool:
        return path.endswith((".xlsx", ".xls"))

    def parse(
        self,
        path: str,
        *,
        station: str = "",
        channel: str = "",
        sheet_name: str | int = 0,
        multi_sheet: bool | None = None,
        y_col: str | int = 0,
        z_col: str | int = 1,
        section_col: str | int | None = None,
        skip_rows: int = 0,
        header_row: bool = True,
        **kwargs: Any,
    ) -> list[SectionProfile]:
        """解析 Excel 断面文件。

        multi_sheet=True: 每个 sheet 是一个独立断面（如 `1#`, `2#`）
        multi_sheet=None: 自动检测（>1 sheet 且 sheet 名匹配断面模式则用多 sheet）
        multi_sheet=False: 单 sheet 内按行分段
        """
        try:
            import openpyxl
        except ImportError:
            try:
                import pandas as pd
                return self._parse_with_pandas(
                    path, station, channel, sheet_name, y_col, z_col, section_col, skip_rows,
                )
            except ImportError:
                raise ImportError("需要 openpyxl 或 pandas 来解析 xlsx 文件")

        if multi_sheet is None:
            multi_sheet = self._should_use_multi_sheet(path)

        if multi_sheet:
            return self._parse_multi_sheet(path, station, channel, y_col, z_col, header_row)

        return self._parse_with_openpyxl(
            path, station, channel, sheet_name, y_col, z_col, section_col, skip_rows,
        )

    def _parse_with_openpyxl(
        self, path: str, station: str, channel: str,
        sheet_name: str | int, y_col: str | int, z_col: str | int,
        section_col: str | int | None, skip_rows: int,
    ) -> list[SectionProfile]:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]

        y_idx = self._resolve_col(y_col)
        z_idx = self._resolve_col(z_col)
        sec_idx = self._resolve_col(section_col) if section_col is not None else None

        sections: dict[str, list[list[float]]] = {}
        current_section = "default"

        for i, row in enumerate(ws.iter_rows(min_row=skip_rows + 1, values_only=True)):
            if sec_idx is not None and row[sec_idx] is not None:
                val = str(row[sec_idx]).strip()
                if val:
                    current_section = val

            try:
                y_val = float(row[y_idx])
                z_val = float(row[z_idx])
                sections.setdefault(current_section, []).append([y_val, z_val])
            except (TypeError, ValueError, IndexError):
                continue

        wb.close()
        return self._to_profiles(sections, path, station, channel)

    def _parse_with_pandas(
        self, path: str, station: str, channel: str,
        sheet_name: str | int, y_col: str | int, z_col: str | int,
        section_col: str | int | None, skip_rows: int,
    ) -> list[SectionProfile]:
        import pandas as pd
        df = pd.read_excel(path, sheet_name=sheet_name, skiprows=skip_rows)

        if isinstance(y_col, str):
            y_series = df[y_col]
        else:
            y_series = df.iloc[:, y_col]

        if isinstance(z_col, str):
            z_series = df[z_col]
        else:
            z_series = df.iloc[:, z_col]

        sections: dict[str, list[list[float]]] = {}
        current_section = "default"

        for i in range(len(df)):
            if section_col is not None:
                val = df.iloc[i, section_col] if isinstance(section_col, int) else df[section_col].iloc[i]
                if pd.notna(val) and str(val).strip():
                    current_section = str(val).strip()
            try:
                y_val = float(y_series.iloc[i])
                z_val = float(z_series.iloc[i])
                sections.setdefault(current_section, []).append([y_val, z_val])
            except (TypeError, ValueError):
                continue

        return self._to_profiles(sections, path, station, channel)

    @staticmethod
    def _to_profiles(
        sections: dict[str, list[list[float]]], path: str, station: str, channel: str,
    ) -> list[SectionProfile]:
        results = []
        for idx, (name, yz) in enumerate(sections.items()):
            if len(yz) < 2:
                continue
            results.append(SectionProfile(
                id=f"{station}_{name}" if station else name,
                name=name,
                yz=yz,
                location=float(idx),
                channel=channel,
                station=station,
                source_type="xlsx_terrain",
                source_path=path,
            ))
        return results

    @staticmethod
    def _should_use_multi_sheet(path: str) -> bool:
        """自动检测：如果有多个 sheet 且名称像断面编号则用多 sheet 模式。"""
        import re
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            names = wb.sheetnames
            wb.close()
        except Exception:
            return False
        if len(names) <= 1:
            return False
        section_pattern = re.compile(r"^\d+#?$|^[A-Z]{1,3}\d+$|^断面", re.IGNORECASE)
        matches = sum(1 for n in names if section_pattern.match(n.strip()))
        return matches > len(names) * 0.5

    def _parse_multi_sheet(
        self, path: str, station: str, channel: str,
        y_col: str | int, z_col: str | int, header_row: bool,
    ) -> list[SectionProfile]:
        """每个 sheet 是一个独立断面。"""
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        y_idx = self._resolve_col(y_col)
        z_idx = self._resolve_col(z_col)

        results: list[SectionProfile] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            yz: list[list[float]] = []
            start_row = 2 if header_row else 1
            for row in ws.iter_rows(min_row=start_row, values_only=True):
                try:
                    y_val = float(row[y_idx])
                    z_val = float(row[z_idx])
                    yz.append([y_val, z_val])
                except (TypeError, ValueError, IndexError):
                    continue
            if len(yz) >= 2:
                results.append(SectionProfile(
                    id=f"{station}_{sheet_name}" if station else sheet_name,
                    name=sheet_name,
                    yz=yz,
                    location=float(idx),
                    channel=channel,
                    station=station,
                    source_type="xlsx_terrain",
                    source_path=path,
                ))

        wb.close()
        return results

    @staticmethod
    def _resolve_col(col: str | int | None) -> int:
        if col is None or isinstance(col, int):
            return col or 0
        return ord(col.upper()) - ord("A")
