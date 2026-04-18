"""Miners for the Hydraulic domain (E1-E5).

Handles cross-sections, Manning roughness, Z-V curves, Z-Q curves,
and flow curves.  Primary sources: wxq JSON, section TXT, CSV, XLSX.
"""
from __future__ import annotations

import csv
import fnmatch
import json
import logging
import re
from pathlib import Path
from typing import Any

from ..registry import MineResult
from ..taxonomy import TYPE_CATALOG, DataType

log = logging.getLogger(__name__)

_HYD_TYPES = [
    DataType.CROSS_SECTION,
    DataType.MANNING_ROUGHNESS,
    DataType.ZV_CURVE,
    DataType.ZQ_CURVE,
    DataType.FLOW_CURVE,
]


def _find_at(obj: Any, key: str) -> Any:
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        for v in obj.values():
            r = _find_at(v, key)
            if r is not None:
                return r
    return None


class HydraulicMiner:
    @property
    def handled_types(self) -> list[DataType]:
        return list(_HYD_TYPES)

    def probe(self, path: Path, cfg: dict[str, Any]) -> list[DataType]:
        matched: list[DataType] = []
        name_lower = path.name.lower()
        ext = path.suffix.lower()
        for dt in _HYD_TYPES:
            meta = TYPE_CATALOG[dt]
            if ext not in meta.extensions:
                continue
            if meta.filename_patterns and not any(
                fnmatch.fnmatch(name_lower, p) for p in meta.filename_patterns
            ):
                if ext == ".json":
                    matched.append(dt)
                    continue
                continue
            matched.append(dt)
        return matched

    def extract(
        self, path: Path, data_type: DataType, cfg: dict[str, Any],
    ) -> list[MineResult]:
        ext = path.suffix.lower()
        if ext == ".json":
            return self._extract_json(path, data_type)
        if ext == ".txt":
            return self._extract_txt(path, data_type)
        if ext == ".csv":
            return self._extract_csv(path, data_type)
        if ext in (".xlsx", ".xls"):
            return self._extract_excel(path, data_type)
        return []

    # ── JSON extraction (wxq topology) ────────────────────────────────────

    def _extract_json(self, path: Path, data_type: DataType) -> list[MineResult]:
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return []

        if data_type == DataType.CROSS_SECTION:
            return self._sections_from_json(path, data)
        if data_type == DataType.MANNING_ROUGHNESS:
            return self._roughness_from_json(path, data)
        if data_type == DataType.ZV_CURVE:
            return self._zv_from_json(path, data)
        if data_type == DataType.ZQ_CURVE:
            return self._zq_from_json(path, data)
        if data_type == DataType.FLOW_CURVE:
            return self._flow_from_json(path, data)
        return []

    def _sections_from_json(self, path: Path, data: Any) -> list[MineResult]:
        base = _find_at(data, "baseData") or {}
        sections = base.get("sections") or _find_at(data, "sections") or {}
        if not isinstance(sections, dict) or not sections:
            return []

        results: list[MineResult] = []
        for name, sec in sections.items():
            yz = sec.get("yz", [])
            if not yz or not isinstance(yz, list):
                continue
            n_points = len(yz)
            z_vals = [p[1] for p in yz if isinstance(p, (list, tuple)) and len(p) >= 2]
            results.append(MineResult(
                data_type=DataType.CROSS_SECTION,
                source_path=str(path),
                source_kind="json_topology",
                payload={
                    "name": name,
                    "n_points": n_points,
                    "data_points": n_points,
                    "z_min": min(z_vals) if z_vals else None,
                    "z_max": max(z_vals) if z_vals else None,
                    "width": (yz[-1][0] - yz[0][0]) if len(yz) >= 2 else None,
                },
                confidence=0.9 if n_points >= 3 else 0.4,
                label=f"断面: {name} ({n_points} pts)",
            ))

        if results:
            return [MineResult(
                data_type=DataType.CROSS_SECTION,
                source_path=str(path),
                source_kind="json_topology",
                payload={
                    "section_count": len(results),
                    "data_points": sum(
                        r.payload.get("n_points", 0) for r in results
                    ),
                    "sections": [r.payload for r in results],
                },
                confidence=0.9,
                label=f"断面集合: {len(results)} sections from {path.name}",
            )]
        return []

    def _roughness_from_json(self, path: Path, data: Any) -> list[MineResult]:
        base = _find_at(data, "baseData") or {}
        channels = base.get("channels") or _find_at(data, "channels") or {}
        if not isinstance(channels, dict) or not channels:
            return []

        items: list[dict] = []
        for name, ch in channels.items():
            nc = ch.get("nc")
            if nc is not None:
                items.append({"channel": name, "manning_n": nc})

        if not items:
            return []
        return [MineResult(
            data_type=DataType.MANNING_ROUGHNESS,
            source_path=str(path),
            source_kind="json_topology",
            payload={
                "roughness_values": items,
                "count": len(items),
                "data_points": len(items),
            },
            confidence=0.8,
            label=f"糙率: {len(items)} channels",
        )]

    def _zv_from_json(self, path: Path, data: Any) -> list[MineResult]:
        base = _find_at(data, "baseData") or {}
        nodes = base.get("nodes") or _find_at(data, "nodes") or {}
        if not isinstance(nodes, dict):
            return []
        results: list[MineResult] = []
        for name, node in nodes.items():
            zv = node.get("ZV") or node.get("zv")
            if not zv or not isinstance(zv, list):
                continue
            results.append(MineResult(
                data_type=DataType.ZV_CURVE,
                source_path=str(path),
                source_kind="json_topology",
                payload={
                    "name": name,
                    "n_points": len(zv),
                    "data_points": len(zv),
                    "curve_data": zv[:5] if len(zv) > 5 else zv,
                },
                confidence=0.9 if len(zv) >= 3 else 0.5,
                label=f"Z-V曲线: {name} ({len(zv)} pts)",
            ))
        return results

    def _zq_from_json(self, path: Path, data: Any) -> list[MineResult]:
        base = _find_at(data, "baseData") or {}
        nodes = base.get("nodes") or _find_at(data, "nodes") or {}
        if not isinstance(nodes, dict):
            return []
        results: list[MineResult] = []
        for name, node in nodes.items():
            zq = node.get("ZQ") or node.get("zq")
            if not zq or not isinstance(zq, list):
                continue
            results.append(MineResult(
                data_type=DataType.ZQ_CURVE,
                source_path=str(path),
                source_kind="json_topology",
                payload={
                    "name": name,
                    "n_points": len(zq),
                    "data_points": len(zq),
                    "curve_data": zq[:5] if len(zq) > 5 else zq,
                },
                confidence=0.9 if len(zq) >= 3 else 0.5,
                label=f"Z-Q曲线: {name} ({len(zq)} pts)",
            ))
        return results

    def _flow_from_json(self, path: Path, data: Any) -> list[MineResult]:
        base = _find_at(data, "baseData") or {}
        gates = base.get("gates") or {}
        if not isinstance(gates, dict):
            return []
        results: list[MineResult] = []
        for name, gdata in gates.items():
            if not isinstance(gdata, dict):
                continue
            has_flow_params = any(
                gdata.get(k) is not None for k in ("c1", "c2", "c3", "c4")
            )
            if not has_flow_params:
                continue
            results.append(MineResult(
                data_type=DataType.FLOW_CURVE,
                source_path=str(path),
                source_kind="json_topology",
                payload={
                    "name": name,
                    "type": "gate_flow",
                    "c1": gdata.get("c1"), "c2": gdata.get("c2"),
                    "c3": gdata.get("c3"), "c4": gdata.get("c4"),
                    "zb": gdata.get("zb"),
                    "width_b": gdata.get("b"),
                    "data_points": sum(
                        1 for k in ("c1", "c2", "c3", "c4")
                        if gdata.get(k) is not None
                    ),
                },
                confidence=0.8,
                label=f"过流曲线: {name}",
            ))
        return results

    # ── TXT extraction (cross-section / terrain files) ────────────────────

    def _extract_txt(self, path: Path, data_type: DataType) -> list[MineResult]:
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            return []

        if data_type == DataType.CROSS_SECTION:
            sections = self._parse_section_txt(text, path)
            if not sections:
                return []
            return [MineResult(
                data_type=DataType.CROSS_SECTION,
                source_path=str(path),
                source_kind="text",
                payload={
                    "section_count": len(sections),
                    "data_points": sum(s.get("n_points", 0) for s in sections),
                    "sections": sections,
                },
                confidence=0.7 if len(sections) >= 1 else 0.3,
                label=f"断面(TXT): {len(sections)} sections from {path.name}",
            )]

        if data_type in (DataType.ZV_CURVE, DataType.ZQ_CURVE):
            return self._curve_from_txt(text, path, data_type)

        return []

    def _curve_from_txt(
        self, text: str, path: Path, data_type: DataType,
    ) -> list[MineResult]:
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        curve_data: list[list[float]] = []
        for line in lines:
            nums = re.findall(r"[-+]?\d*\.?\d+", line)
            if len(nums) >= 2:
                try:
                    curve_data.append([float(nums[0]), float(nums[1])])
                except ValueError:
                    pass
        if len(curve_data) < 2:
            return []
        return [MineResult(
            data_type=data_type,
            source_path=str(path),
            source_kind="text",
            payload={
                "name": path.stem,
                "n_points": len(curve_data),
                "data_points": len(curve_data),
                "curve_data": curve_data[:5],
            },
            confidence=0.7 if len(curve_data) >= 5 else 0.4,
            label=f"{TYPE_CATALOG[data_type].label_cn}: {path.stem} ({len(curve_data)} pts)",
        )]

    def _parse_section_txt(self, text: str, path: Path) -> list[dict]:
        """Parse common cross-section TXT formats."""
        sections: list[dict] = []
        lines = text.splitlines()
        current_name: str | None = None
        current_points: list[list[float]] = []

        for line in lines:
            stripped = line.strip()
            if not stripped:
                continue
            sec_match = re.match(r"断面\s*(\S+)", stripped)
            if sec_match or re.match(r"^[^\d\s.-]+", stripped):
                if current_name and current_points:
                    sections.append(self._section_summary(current_name, current_points))
                current_name = sec_match.group(1) if sec_match else stripped
                current_points = []
                continue
            nums = re.findall(r"[-+]?\d*\.?\d+", stripped)
            if len(nums) >= 2:
                try:
                    current_points.append([float(nums[0]), float(nums[1])])
                except ValueError:
                    pass

        if current_name and current_points:
            sections.append(self._section_summary(current_name, current_points))
        return sections

    @staticmethod
    def _section_summary(name: str, points: list[list[float]]) -> dict:
        z_vals = [p[1] for p in points]
        return {
            "name": name,
            "n_points": len(points),
            "z_min": min(z_vals) if z_vals else None,
            "z_max": max(z_vals) if z_vals else None,
            "width": abs(points[-1][0] - points[0][0]) if len(points) >= 2 else None,
        }

    # ── Excel extraction (curves, sections) ─────────────────────────────

    def _extract_excel(self, path: Path, data_type: DataType) -> list[MineResult]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except ImportError:
            return []
        except Exception:
            return []

        results: list[MineResult] = []

        if data_type == DataType.CROSS_SECTION:
            total_points = 0
            sheet_count = 0
            for ws in wb.worksheets:
                rows = sum(1 for _ in ws.iter_rows(min_row=2, max_col=2, values_only=True))
                if rows > 0:
                    total_points += rows
                    sheet_count += 1
            if total_points > 0:
                results.append(MineResult(
                    data_type=data_type,
                    source_path=str(path),
                    source_kind="xlsx",
                    payload={
                        "section_count": sheet_count,
                        "data_points": total_points,
                    },
                    confidence=0.8,
                    label=f"断面(XLSX): {path.name} ({sheet_count}页/{total_points}点)",
                ))

        elif data_type in (DataType.ZV_CURVE, DataType.ZQ_CURVE):
            for ws in wb.worksheets:
                curve_data: list[list[float]] = []
                for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                    if row[0] is not None and row[1] is not None:
                        try:
                            curve_data.append([float(row[0]), float(row[1])])
                        except (ValueError, TypeError):
                            pass
                if len(curve_data) >= 2:
                    results.append(MineResult(
                        data_type=data_type,
                        source_path=str(path),
                        source_kind="xlsx",
                        payload={
                            "name": f"{path.stem}/{ws.title}",
                            "n_points": len(curve_data),
                            "data_points": len(curve_data),
                            "curve_data": curve_data[:5],
                        },
                        confidence=0.8 if len(curve_data) >= 5 else 0.5,
                        label=f"{TYPE_CATALOG[data_type].label_cn}: {path.stem}/{ws.title}",
                    ))

        wb.close()
        return results

    # ── CSV extraction (curves) ───────────────────────────────────────────

    def _extract_csv(self, path: Path, data_type: DataType) -> list[MineResult]:
        try:
            text = path.read_text(encoding="utf-8", errors="replace")
            reader = csv.DictReader(text.splitlines())
            fields = reader.fieldnames or []
        except Exception:
            return []

        rows = list(reader)
        if not rows:
            return []

        if data_type in (DataType.ZV_CURVE, DataType.ZQ_CURVE):
            z_cols = [k for k in fields if any(w in k.lower() for w in ["z", "水位", "level"])]
            v_cols = [k for k in fields if any(w in k.lower() for w in ["v", "容", "q", "流量", "storage"])]
            if z_cols and v_cols:
                curve_data: list[list[float]] = []
                for row in rows:
                    try:
                        curve_data.append([float(row[z_cols[0]]), float(row[v_cols[0]])])
                    except (ValueError, KeyError):
                        pass
                if curve_data:
                    return [MineResult(
                        data_type=data_type,
                        source_path=str(path),
                        source_kind="csv",
                        payload={
                            "name": path.stem,
                            "n_points": len(curve_data),
                            "data_points": len(curve_data),
                            "curve_data": curve_data[:5],
                        },
                        confidence=0.7 if len(curve_data) >= 3 else 0.4,
                        label=f"{TYPE_CATALOG[data_type].label_cn}: {path.stem}",
                    )]
        return []
